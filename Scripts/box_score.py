#!/usr/bin/env python

"""box_score.py: Parse eggball NDJSON records and print a single box score table.

All players are auto-detected from the NDJSON files — no player list needed.
Rows are sorted by Bad% ascending (best player first).

Usage:
    python box_score.py "<NDJSON_DIR>"

Example:
    python box_score.py "NDJSONs/2024-04-17/"
"""

import csv
import ndjson
import os
import sys
import tabulate
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUTPUT_DIR = Path(__file__).parent.parent / 'Box Scores'

# Players to exclude from all output. Add names (lowercase) to hide them.
EXCLUDE_PLAYERS = {
    # 'blaster',
}

DISPLAY_HEADERS = [
    'Player', 'Hold (secs)', 'Egg Start', 'Receptions', 'Ints Caught', 'Tags',
    'Total Touches', 'Time Held PP', 'Self Passes', 'Touches w/ SP',
    'Completions', 'Caps', 'Raps Thrown', 'Raps Caught',
    'Good', 'Good %', 'Ints Thrown', 'Tagged', 'Bad', 'Bad %'
]
# Indices averaged (not summed) in the totals row
_AVG_FLOAT_IDX = {7}    # Hold Time
_AVG_PCT_IDX   = {15, 19}  # Good %, Bad %

RAW_KEYS = [
    'hold', 'start_egg', 'receptions', 'self_passes', 'ints_caught', 'tags',
    'caps', 'raps_thrown', 'completions', 'ints_thrown', 'tagged', 'raps_caught'
]


def player_row(name, p):
    touches = p['start_egg'] + p['receptions'] + p['ints_caught'] + p['tags']
    hold_time = round(p['hold'] / touches, 2) if touches else 0.0
    touches_wsp = touches + p['self_passes']
    good = p['completions'] + p['caps'] + p['raps_thrown'] + p['raps_caught']
    good_pct = f"{good / touches * 100:.1f}%" if touches else "0.0%"
    bad = p['ints_thrown'] + p['tagged']
    bad_pct = f"{bad / touches * 100:.1f}%" if touches else "0.0%"
    return [
        name, round(p['hold'], 1),
        p['start_egg'], p['receptions'], p['ints_caught'], p['tags'],
        touches, hold_time, p['self_passes'], touches_wsp,
        p['completions'], p['caps'], p['raps_thrown'], p['raps_caught'],
        good, good_pct,
        p['ints_thrown'], p['tagged'], bad, bad_pct
    ]


def totals_row(rows):
    n = len(rows)
    result = ['Total']
    for i in range(1, len(DISPLAY_HEADERS)):
        col_vals = [r[i] for r in rows]
        if i in _AVG_FLOAT_IDX:
            result.append(round(sum(col_vals) / n, 2))
        elif i in _AVG_PCT_IDX:
            nums = [float(v.rstrip('%')) for v in col_vals]
            result.append(f"{sum(nums) / n:.1f}%")
        else:
            result.append(sum(col_vals))
    return result


def _should_exclude(name):
    lower = name.lower()
    return lower in EXCLUDE_PLAYERS or lower.startswith('some ball')


def build_table(final):
    """Return sorted player rows + totals row for all players in final."""
    rows = [player_row(name, p) for name, p in final.items()
            if not _should_exclude(name)]
    rows.sort(key=lambda r: float(r[19].rstrip('%')))
    if rows:
        rows.append(totals_row(rows))
    return rows


def build_per_game_table(final, n_games):
    """Return a per-game stats table: each player's raw stats divided by n_games.
    Percentages and Hold Time are rates and stay unchanged."""
    keep_same = _AVG_FLOAT_IDX | _AVG_PCT_IDX  # indices 7, 15, 19

    def pg_player_row(name, p):
        base = player_row(name, p)
        row = [name]
        for i in range(1, len(DISPLAY_HEADERS)):
            val = base[i]
            if i in keep_same:
                row.append(val)
            else:
                row.append(round(val / n_games, 1))
        return row

    rows = [pg_player_row(name, p) for name, p in final.items()
            if not _should_exclude(name)]
    rows.sort(key=lambda r: float(r[19].rstrip('%')))
    return rows


def infer_game_format(data):
    """Return '5v5', '6v6', or 'other' based on roster size."""
    if data:
        first = data[0]
        if (
            isinstance(first, list)
            and len(first) >= 3
            and first[1] == 'recorder-metadata'
            and isinstance(first[2], dict)
        ):
            players = first[2].get('players')
            if isinstance(players, list):
                team_counts = defaultdict(int)
                for player in players:
                    team = player.get('team')
                    if team in (1, 2):
                        team_counts[team] += 1
                if team_counts[1] == 5 and team_counts[2] == 5:
                    return '5v5'
                if team_counts[1] == 6 and team_counts[2] == 6:
                    return '6v6'

    for line in data:
        if line[1] == 'p' and isinstance(line[2], list) and line[2] and 'name' in line[2][0]:
            team_counts = defaultdict(int)
            for player in line[2]:
                team = player.get('team')
                if team in (1, 2):
                    team_counts[team] += 1
            if team_counts[1] == 5 and team_counts[2] == 5:
                return '5v5'
            if team_counts[1] == 6 and team_counts[2] == 6:
                return '6v6'
            break

    return 'other'


def append_table_section(sections, csv_rows, title, subtitle, final, n_games, per_game=False):
    if not final or n_games <= 0:
        return

    table = build_per_game_table(final, n_games) if per_game else build_table(final)
    if not table:
        return

    sections.append(f"\n\n{'=' * 60}")
    sections.append(f"  {title}  ({subtitle})")
    sections.append(f"{'=' * 60}\n")
    sections.append(tabulate.tabulate(table, DISPLAY_HEADERS))

    csv_rows.append([])
    csv_rows.append([f"{title} ({subtitle})"])
    csv_rows.append(DISPLAY_HEADERS)
    csv_rows.extend(table)


def _coerce_excel_value(value):
    if isinstance(value, str) and value.endswith('%'):
        try:
            return float(value[:-1]) / 100
        except ValueError:
            return value
    return value


def _apply_number_format(cell):
    if cell.column == 1 or cell.value is None:
        return
    if cell.column == 2:
        cell.number_format = '#,##0.0'
        return
    if cell.column in (16, 20):
        cell.number_format = '0%'
    elif isinstance(cell.value, float):
        cell.number_format = '#,##0.0'
    else:
        cell.number_format = '#,##0'


def _hex_to_rgb(hex_color):
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i + 2], 16) for i in (0, 2, 4))


def _rgb_to_hex(rgb):
    return ''.join(f'{max(0, min(255, int(round(channel)))):02X}' for channel in rgb)


def _blend_colors(color_a, color_b, ratio):
    rgb_a = _hex_to_rgb(color_a)
    rgb_b = _hex_to_rgb(color_b)
    return _rgb_to_hex(
        tuple(rgb_a[i] + (rgb_b[i] - rgb_a[i]) * ratio for i in range(3))
    )


def _apply_value_fills(ws, start_row, end_row, col_idx, reverse=False):
    values = []
    for row_idx in range(start_row, end_row + 1):
        value = ws.cell(row=row_idx, column=col_idx).value
        if isinstance(value, (int, float)):
            values.append((row_idx, float(value)))

    if not values:
        return

    sorted_unique = sorted({value for _, value in values})
    denom = max(len(sorted_unique) - 1, 1)
    ranks = {value: idx / denom for idx, value in enumerate(sorted_unique)}

    for row_idx, value in values:
        ratio = ranks[value] if len(sorted_unique) > 1 else 0.5
        if reverse:
            ratio = 1 - ratio
        if ratio <= 0.2:
            fill_color = 'E6C7C3'
        elif ratio <= 0.4:
            fill_color = 'F1DDD8'
        elif ratio < 0.6:
            fill_color = 'F7F5EF'
        elif ratio < 0.8:
            fill_color = 'DFEBDD'
        else:
            fill_color = 'C8DDC4'
        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(fill_type='solid', fgColor=fill_color)


def _section_highlight_columns(title):
    full_cols = {4, 5, 6, 7, 8, 11, 12, 13, 14, 16, 17, 18, 20}
    return full_cols


def _use_base_fill_for_data(title):
    return True


def write_formatted_xlsx(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Box Score'

    title_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    header_fill = PatternFill(fill_type='solid', fgColor='D9EAF7')
    total_fill = PatternFill(fill_type='solid', fgColor='E2F0D9')
    emphasis_fill = PatternFill(fill_type='solid', fgColor='EEF4EA')
    thin = Side(style='thin', color='A6A6A6')
    current_section_start = None
    current_section_title = ''
    current_section_last_data_row = None
    section_ranges = []

    def finalize_section():
        if (
            current_section_start is not None
            and current_section_last_data_row is not None
            and current_section_start <= current_section_last_data_row
        ):
            section_ranges.append((current_section_start, current_section_last_data_row, current_section_title))

    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_coerce_excel_value(value))
            if len(row) > 1 and row != DISPLAY_HEADERS and row[0] != 'Total':
                _apply_number_format(cell)

        if not row:
            continue

        if len(row) == 1:
            finalize_section()
            current_section_title = str(row[0])
            current_section_start = row_idx + 2
            current_section_last_data_row = None
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(DISPLAY_HEADERS))
            cell = ws.cell(row=row_idx, column=1)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = title_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row_idx].height = 22
            continue

        for cell in ws[row_idx]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

        if row == DISPLAY_HEADERS:
            emphasized_cols = _section_highlight_columns(current_section_title)
            for cell in ws[row_idx]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[row_idx].height = 34
        elif row[0] == 'Total':
            emphasized_cols = _section_highlight_columns(current_section_title)
            for cell in ws[row_idx]:
                cell.font = Font(bold=True)
                cell.fill = total_fill
                _apply_number_format(cell)
            if _use_base_fill_for_data(current_section_title):
                for col_idx in emphasized_cols:
                    ws.cell(row=row_idx, column=col_idx).fill = emphasis_fill
        else:
            emphasized_cols = _section_highlight_columns(current_section_title)
            if _use_base_fill_for_data(current_section_title):
                for col_idx in emphasized_cols:
                    ws.cell(row=row_idx, column=col_idx).fill = emphasis_fill
            current_section_last_data_row = row_idx

    widths = {
        1: 14, 2: 10, 3: 8, 4: 8, 5: 8, 6: 7, 7: 10, 8: 9, 9: 8, 10: 10,
        11: 9, 12: 6, 13: 8, 14: 8, 15: 6, 16: 7, 17: 8, 18: 7, 19: 6, 20: 7
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    finalize_section()

    for start_row, end_row, section_title in section_ranges:
        emphasized_cols = _section_highlight_columns(section_title)
        for col_idx in emphasized_cols:
            if col_idx in (8, 17, 18, 20):
                continue
            if col_idx == 16:
                _apply_value_fills(ws, start_row, end_row, col_idx)
            else:
                _apply_value_fills(ws, start_row, end_row, col_idx)
        for col_idx in emphasized_cols:
            if col_idx in (8, 17, 18, 20):
                _apply_value_fills(ws, start_row, end_row, col_idx, reverse=True)

    wb.save(path)


class BoxScore:
    def __init__(self, input_dir, auto_produce=True):
        self.dir = input_dir
        if auto_produce:
            self.produce_stats()

    def produce_stats(self):
        records = []
        records_by_format = defaultdict(list)
        games_by_format = defaultdict(int)
        processed_games = 0
        for filename in os.listdir(self.dir):
            if not filename.endswith('.ndjson'):
                continue
            path = Path(self.dir) / filename
            with open(path) as f:
                try:
                    data = ndjson.load(f)
                    results = self.one_game(data)
                    game_format = infer_game_format(data)
                    records.extend(results)
                    records_by_format[game_format].extend(results)
                    games_by_format[game_format] += 1
                    processed_games += 1
                except ValueError:
                    continue
        final = self.merge_results(records)
        self.print_box_score(final, processed_games, records_by_format, games_by_format)

    def one_game(self, data):
        ids = self.initialize_game(data)
        eggball = [x for x in data if x[1] == 'eggBall' or x[1] == 'boat']
        eggball = [x for x in eggball if x[1] == 'boat' or x[2]['state']]
        eggball = eggball[1:]
        huddleSwitch = False
        prevHolder = {}
        prevPrevHolder = {}
        holder = {}
        caughtAt = 0
        in_air = True
        waiting_last = False
        for i in eggball:
            timestamp = i[0]
            dic = i[2]
            if i[1] == 'boat':
                if prevPrevHolder.get('team') == prevHolder.get('team'):
                    prevPrevHolder['raps_thrown'] += 1
                    prevPrevHolder['caps'] += 1
                    prevHolder['raps_caught'] += 1
                else:
                    prevHolder['raps_caught'] += 1
                waiting_last = False
                continue
            if dic['state'] == 'waiting':
                if waiting_last:
                    continue
                try:
                    prevHolder['caps'] += 1
                    prevHolder['hold'] += ((timestamp - caughtAt) / 1000)
                except KeyError:
                    # sys.stderr.write('No prevHolder ' + str(timestamp))
                    pass
                waiting_last = True
                continue
            waiting_last = False
            holderid = dic['holder']
            if holderid:
                holder = ids[holderid]
            if dic['state'] == 'huddle':
                if huddleSwitch:
                    holder['start_egg'] += 1
                    prevHolder = ids[holderid]
                    caughtAt = timestamp + 5000
                huddleSwitch = not huddleSwitch
            elif not in_air:
                if holderid and holderid != prevHolder.get('id'):
                    if holder['team'] == prevHolder.get('team'):
                        holder['receptions'] += 1
                        prevHolder['completions'] += 1
                    else:
                        holder['tags'] += 1
                        prevHolder['tagged'] += 1
                prevHolder['hold'] += ((timestamp - caughtAt) / 1000)
                caughtAt = timestamp
            elif holderid:
                if holder['team'] != prevHolder.get('team'):
                    prevHolder['ints_thrown'] += 1
                    holder['ints_caught'] += 1
                elif holder != prevHolder:
                    prevHolder['completions'] += 1
                    holder['receptions'] += 1
                else:
                    holder['self_passes'] += 1
                caughtAt = timestamp
            else:
                prevHolder['hold'] += ((timestamp - caughtAt) / 1000)
            if holderid:
                prevPrevHolder = prevHolder
                prevHolder = ids[holderid]
                in_air = False
            else:
                in_air = True
        return ids[1:]

    def initialize_game(self, data):
        ids = [None] * 50
        ids[0] = {'player': 'Brazzers', 'team': 0}
        start_time = 0
        end_time = -1
        for line in data:
            if line[1] == 'p' and type(line[2]) is list and 'name' in line[2][0]:
                for player in line[2]:
                    id = player['id']
                    ids[id] = {
                        'id': id,
                        'player': player['name'],
                        'team': player['team'],
                        'hold': 0.0,
                        'start_egg': 0,
                        'receptions': 0,
                        'self_passes': 0,
                        'ints_caught': 0,
                        'tags': 0,
                        'caps': 0,
                        'raps_thrown': 0,
                        'completions': 0,
                        'ints_thrown': 0,
                        'tagged': 0,
                        'raps_caught': 0,
                        'joined_at': line[0],
                        'left_at': -1
                    }
            elif line[1] == 'time' and line[2]['state'] == 1:
                start_time = line[0]
            elif line[1] == 'playerLeft':
                if end_time == -1:
                    ids[line[2]]['left_at'] = line[0]
            elif line[1] == 'end':
                end_time = line[0]
        ids = [i for i in ids if i]
        for i in ids[1:]:
            i['joined_at'] = max(i['joined_at'], start_time)
            if i['left_at'] == -1:
                i['left_at'] = end_time
        return ids

    def merge_results(self, records):
        final = {}
        for r in records:
            player = r['player']
            if player in final:
                for k in RAW_KEYS:
                    final[player][k] += r[k]
            else:
                final[player] = {k: r[k] for k in RAW_KEYS}
                final[player]['team'] = r['team']
        return final

    def print_box_score(self, final, total_games, records_by_format, games_by_format):
        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        OUTPUT_DIR.mkdir(exist_ok=True)
        sections = []
        csv_rows = []

        append_table_section(
            sections,
            csv_rows,
            'BOX SCORE TOTALS',
            f"{total_games} games",
            final,
            total_games,
        )
        append_table_section(
            sections,
            csv_rows,
            'BOX SCORE PER GAME',
            f"{total_games} games",
            final,
            total_games,
            per_game=True,
        )
        append_table_section(
            sections,
            csv_rows,
            'BOX SCORE 6V6',
            f"{games_by_format['6v6']} games",
            self.merge_results(records_by_format['6v6']),
            games_by_format['6v6'],
        )
        append_table_section(
            sections,
            csv_rows,
            'BOX SCORE 6V6 PER GAME',
            f"{games_by_format['6v6']} games",
            self.merge_results(records_by_format['6v6']),
            games_by_format['6v6'],
            per_game=True,
        )
        append_table_section(
            sections,
            csv_rows,
            'BOX SCORE 5V5',
            f"{games_by_format['5v5']} games",
            self.merge_results(records_by_format['5v5']),
            games_by_format['5v5'],
        )
        append_table_section(
            sections,
            csv_rows,
            'BOX SCORE 5V5 PER GAME',
            f"{games_by_format['5v5']} games",
            self.merge_results(records_by_format['5v5']),
            games_by_format['5v5'],
            per_game=True,
        )
        
        # Write TXT
        path = OUTPUT_DIR / f'box_score_{timestamp}.txt'
        with open(path, 'w') as f:
            f.write('\n'.join(sections))
        print(f"Written to {path}")

        # Write CSV
        csv_path = OUTPUT_DIR / f'box_score_{timestamp}.csv'
        with open(csv_path, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(csv_rows)
        print(f"Written to {csv_path}")

        xlsx_path = OUTPUT_DIR / f'box_score_{timestamp}.xlsx'
        write_formatted_xlsx(xlsx_path, csv_rows)
        print(f"Written to {xlsx_path}")


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print("Usage: python box_score.py <ndjson_dir>")
        sys.exit(1)
    BoxScore(sys.argv[1])
